import openpyxl
from collections import defaultdict
import json
import re

def determinar_area(programa):
    """Determina el área de conocimiento basado en el programa"""
    programa_lower = str(programa).lower().strip()
    
    # Económico Administrativa
    economico_admin = [
        "administración", "negocios", "mercadotecnia", "marketing", "contaduría", 
        "contabilidad", "finanzas", "comercio", "turismo", "hotelería", 
        "recursos humanos", "economía", "comercial", "empresarial", "gestión"
    ]
    
    # Ciencias de la salud
    ciencias_salud = [
        "medicina", "enfermería", "psicología", "nutrición", "odontología", 
        "dentista", "farmacia", "farmacéutico", "fisioterapia", "terapia", 
        "rehabilitación", "optometría", "veterinaria", "biomédica", "salud"
    ]
    
    # Arquitectura y diseño
    arquitectura_diseño = [
        "arquitectura", "diseño", "gráfico", "industrial", "interiores", 
        "urbanismo", "planeación urbana", "arte", "bellas artes", "visual"
    ]
    
    # Gastronomía
    gastronomia = [
        "gastronomía", "culinaria", "chef", "cocina", "alimentos", "bebidas"
    ]
    
    # Ciencias Sociales y Humanidades
    sociales_humanidades = [
        "derecho", "leyes", "jurisprudencia", "criminología", "criminalística", 
        "trabajo social", "comunicación", "periodismo", "filosofía", "historia", 
        "sociología", "antropología", "geografía", "educación", "pedagogía", 
        "literatura", "idiomas", "lenguas", "relaciones internacionales", 
        "ciencias políticas", "humanidades"
    ]
    
    # Ciencias exactas e ingenierías
    exactas_ingenierias = [
        "ingeniería", "sistemas", "informática", "computación", "software", 
        "matemáticas", "física", "química", "biología", "biotecnología", 
        "industrial", "civil", "mecánica", "eléctrica", "electrónica", 
        "telecomunicaciones", "ambiental", "agronomía", "forestal", 
        "ciencias", "tecnología", "robótica", "inteligencia artificial"
    ]
    
    # Buscar coincidencias
    for palabra in economico_admin:
        if palabra in programa_lower:
            return "Económico Administrativa"
    
    for palabra in ciencias_salud:
        if palabra in programa_lower:
            return "Ciencias de la salud"
    
    for palabra in arquitectura_diseño:
        if palabra in programa_lower:
            return "Arquitectura y diseño"
    
    for palabra in gastronomia:
        if palabra in programa_lower:
            return "Gastronomía"
    
    for palabra in sociales_humanidades:
        if palabra in programa_lower:
            return "Ciencias Sociales y Humanidades"
    
    for palabra in exactas_ingenierias:
        if palabra in programa_lower:
            return "Ciencias exactas e ingenierías"
    
    return "Sin clasificar"

def determinar_nivel_educativo(incorporacion, programa):
    """Determina el nivel educativo basado en la incorporación y programa"""
    
    if not incorporacion:
        incorporacion = ""
    
    incorporacion_lower = str(incorporacion).lower().strip()
    programa_lower = str(programa).lower().strip()
    
    # NIVELES A OMITIR (retornamos None para que se excluyan)
    
    # BIS - Bachillerato Internacional Escolarizado
    if programa_lower == "bis" or "bis" in programa_lower:
        return None
    
    # Maestrías
    maestrias_palabras = ["maestría", "maestrias", "master", "mba", "magister", "posgrado"]
    for palabra in maestrias_palabras:
        if palabra in incorporacion_lower or palabra in programa_lower:
            return None
    
    # Bachillerato y niveles básicos
    if any(word in incorporacion_lower or word in programa_lower 
           for word in ["bachillerato", "preparatoria", "high school", "bachiller", 
                       "primaria", "educación básica", "elementary",
                       "secundaria", "middle school", "educación media"]):
        return None
    
    # Doctorados
    if any(word in incorporacion_lower or word in programa_lower 
           for word in ["doctorado", "phd", "doctor"]):
        return "Doctorados"
    
    # Programas que típicamente son de Educación Continua
    educacion_continua_programas = [
        "diplomado", "curso", "certificación", "capacitación", "continua", 
        "actualización", "especialidad médica", "especialización"
    ]
    for patron in educacion_continua_programas:
        if patron in programa_lower or patron in incorporacion_lower:
            return "Educación Continua"
    
    # CASOS TÉCNICOS
    if any(word in incorporacion_lower or word in programa_lower 
           for word in ["técnico", "tsu", "profesional asociado", "tecnólogo"]):
        return "Técnico Superior"
    
    # LICENCIATURAS - Ahora es mucho más inclusivo
    # Excluir solo casos muy específicos que NO son licenciaturas
    exclusiones_licenciatura = [
        "diplomado"
    ]
    
    # Si no contiene palabras de exclusión y tiene más de 3 caracteres, probablemente es licenciatura
    es_exclusion = any(excl in programa_lower or excl in incorporacion_lower 
                      for excl in exclusiones_licenciatura)
    
    if not es_exclusion and len(programa_lower) > 3:
        return "Licenciaturas"
    
    # Por defecto, sin clasificar
    return "Sin Clasificar"

def procesar_hoja_mejorada(ws, nombre_hoja):
    """Procesa una hoja específica del archivo Excel con mejor detección de estructura"""
    resultados_hoja = defaultdict(list)
    
    plantel_actual = None
    incorporacion_actual = None
    modalidad_actual = None
    
    # Encontrar inicio de los datos
    start_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and len(row) > 0 and row[0] and "Plantel" in str(row[0]):
            start_row = i + 1
            break
    
    if start_row is None:
        print(f"⚠️  No se encontró encabezado 'Plantel' en la hoja {nombre_hoja}")
        return resultados_hoja
    
    print(f"📊 Procesando desde la fila {start_row} en {nombre_hoja}")
    
    # Recorrer desde el inicio real
    for i in range(start_row, ws.max_row + 1):
        try:
            row = [cell.value if cell.value is not None else "" for cell in ws[i]]
            
            # Asegurar que tenemos al menos 4 columnas
            while len(row) < 4:
                row.append("")
            
            # Detectar nuevo plantel (columna A) - el plantel se extiende verticalmente
            if row[0] and str(row[0]).strip() and str(row[0]).strip() not in ["Plantel", ""]:
                plantel_candidato = str(row[0]).strip()
                # Verificar que no sea un programa o dato erróneo
                if len(plantel_candidato) > 2 and not plantel_candidato.lower() in ['udg', 'sicyt']:
                    plantel_actual = plantel_candidato
                    print(f"🏫 Nuevo plantel detectado: {plantel_actual}")
            
            # Detectar nueva incorporación (columna B)
            if row[1] and str(row[1]).strip():
                incorporacion_candidata = str(row[1]).strip()
                # Solo actualizar si parece ser una incorporación válida
                if incorporacion_candidata not in ["", "Incorporación"] and len(incorporacion_candidata) > 2:
                    incorporacion_actual = incorporacion_candidata
                    print(f"📜 Nueva incorporación: {incorporacion_actual}")
            
            # Detectar modalidad (columna D)
            if row[3] and str(row[3]).strip():
                modalidad_candidata = str(row[3]).strip()
                if modalidad_candidata not in ["", "Modalidad"] and len(modalidad_candidata) > 2:
                    modalidad_actual = modalidad_candidata
            
            # Extraer programa (columna C)
            programa = row[2]
            if programa and str(programa).strip() and plantel_actual:
                programa_str = str(programa).strip()
                # Limpiar el programa
                programa_str = ' '.join(programa_str.split())
                
                # Verificar que no sea un encabezado
                if programa_str not in ["Programa", ""] and len(programa_str) > 2:
                    # Determinar nivel educativo y área
                    nivel_educativo = determinar_nivel_educativo(incorporacion_actual, programa_str)
                    
                    # Si el nivel es None (niveles a omitir), saltamos este programa
                    if nivel_educativo is None:
                        print(f"    ⏭️  Programa omitido: '{programa_str}' (nivel excluido)")
                        continue
                    
                    area = determinar_area(programa_str)
                    
                    # Debug: mostrar clasificación para casos problemáticos
                    if not incorporacion_actual or incorporacion_actual == "No especificada":
                        print(f"    🔍 Programa sin incorporación: '{programa_str}' → Nivel: {nivel_educativo}, Área: {area}")
                    
                    programa_info = {
                        "programa": programa_str,
                        "incorporacion": incorporacion_actual or "No especificada",
                        "modalidad": modalidad_actual or "No especificada",
                        "hoja": nombre_hoja,
                        "nivel_educativo": nivel_educativo,
                        "area": area
                    }
                    
                    resultados_hoja[plantel_actual].append(programa_info)
                    print(f"  📚 Programa agregado: {programa_str} ({nivel_educativo} - {area})")
        
        except Exception as e:
            print(f"⚠️  Error procesando fila {i}: {e}")
            continue
    
    return resultados_hoja

def generar_estructura_educativa(todos_resultados):
    """Genera estructura organizada por nivel educativo > área > programa > plantel"""
    estructura = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))
    
    for plantel, programas in todos_resultados.items():
        if not plantel or str(plantel).strip() == "":
            continue
            
        plantel_limpio = str(plantel).strip()
        
        for programa_info in programas:
            programa = programa_info["programa"]
            nivel = programa_info["nivel_educativo"]
            area = programa_info["area"]
            
            # Saltar programas sin clasificar o niveles excluidos
            if not programa or str(programa).strip() == "" or nivel is None:
                continue
            
            programa_limpio = ' '.join(str(programa).split())
            
            # Evitar duplicados usando una clave única
            modalidad_incorporacion = f"{programa_info['modalidad']}|{programa_info['incorporacion']}"
            
            # Verificar si ya existe esta combinación
            existe = False
            for item in estructura[nivel][area][programa_limpio][plantel_limpio]:
                item_key = f"{item['modalidad']}|{item['incorporacion']}"
                if item_key == modalidad_incorporacion:
                    existe = True
                    break
            
            if not existe:
                estructura[nivel][area][programa_limpio][plantel_limpio].append({
                    "modalidad": programa_info["modalidad"],
                    "incorporacion": programa_info["incorporacion"],
                    "hoja_fuente": programa_info["hoja"]
                })
    
    return estructura

def generar_javascript(estructura_final):
    """Genera el archivo JavaScript con los datos"""
    js_content = """// Datos de programas educativos organizados por nivel, área y plantel
// Generado automáticamente desde datos de Excel

export const programasEducativos = """
    
    # Convertir la estructura a un formato más limpio para JS
    js_data = {}
    for nivel in sorted(estructura_final.keys()):
        js_data[nivel] = {}
        for area in sorted(estructura_final[nivel].keys()):
            js_data[nivel][area] = {}
            for programa in sorted(estructura_final[nivel][area].keys()):
                js_data[nivel][area][programa] = {}
                for plantel in sorted(estructura_final[nivel][area][programa].keys()):
                    js_data[nivel][area][programa][plantel] = estructura_final[nivel][area][programa][plantel]
    
    # Agregar los datos JSON al contenido JS
    js_content += json.dumps(js_data, ensure_ascii=False, indent=2)
    js_content += ";\n\n"
    
    # Agregar funciones de utilidad
    js_content += """
// Funciones de utilidad para consultar los datos

export const areas = [
  "Todas",
  "Económico Administrativa",
  "Ciencias de la salud",
  "Arquitectura y diseño",
  "Gastronomía",
  "Ciencias Sociales y Humanidades",
  "Ciencias exactas e ingenierías"
];

export const niveles = Object.keys(programasEducativos);

/**
 * Obtiene todos los programas de un nivel educativo específico
 * @param {string} nivel - Nivel educativo
 * @returns {Object} Programas del nivel especificado
 */
export function obtenerProgramasPorNivel(nivel) {
  return programasEducativos[nivel] || {};
}

/**
 * Obtiene todos los programas de un área específica
 * @param {string} area - Área de conocimiento
 * @param {string} nivel - Nivel educativo (opcional)
 * @returns {Object} Programas del área especificada
 */
export function obtenerProgramasPorArea(area, nivel = null) {
  const resultado = {};
  
  if (nivel) {
    return programasEducativos[nivel]?.[area] || {};
  }
  
  for (const nivelKey in programasEducativos) {
    if (programasEducativos[nivelKey][area]) {
      resultado[nivelKey] = programasEducativos[nivelKey][area];
    }
  }
  
  return resultado;
}

/**
 * Busca programas por nombre (búsqueda parcial)
 * @param {string} termino - Término de búsqueda
 * @returns {Array} Array de programas que coinciden con el término
 */
export function buscarProgramas(termino) {
  const resultados = [];
  const terminoLower = termino.toLowerCase();
  
  for (const nivel in programasEducativos) {
    for (const area in programasEducativos[nivel]) {
      for (const programa in programasEducativos[nivel][area]) {
        if (programa.toLowerCase().includes(terminoLower)) {
          resultados.push({
            programa,
            nivel,
            area,
            planteles: Object.keys(programasEducativos[nivel][area][programa])
          });
        }
      }
    }
  }
  
  return resultados;
}

/**
 * Obtiene información detallada de un programa específico
 * @param {string} programa - Nombre del programa
 * @param {string} plantel - Nombre del plantel
 * @returns {Object|null} Información del programa o null si no existe
 */
export function obtenerDetallePrograma(programa, plantel) {
  for (const nivel in programasEducativos) {
    for (const area in programasEducativos[nivel]) {
      if (programasEducativos[nivel][area][programa]?.[plantel]) {
        return {
          programa,
          nivel,
          area,
          plantel,
          detalles: programasEducativos[nivel][area][programa][plantel]
        };
      }
    }
  }
  return null;
}

/**
 * Obtiene estadísticas generales de los datos
 * @returns {Object} Estadísticas de los programas
 */
export function obtenerEstadisticas() {
  let totalProgramas = 0;
  const planteles = new Set();
  const estadisticasPorNivel = {};
  const estadisticasPorArea = {};
  
  for (const nivel in programasEducativos) {
    estadisticasPorNivel[nivel] = 0;
    
    for (const area in programasEducativos[nivel]) {
      if (!estadisticasPorArea[area]) {
        estadisticasPorArea[area] = 0;
      }
      
      for (const programa in programasEducativos[nivel][area]) {
        totalProgramas++;
        estadisticasPorNivel[nivel]++;
        estadisticasPorArea[area]++;
        
        for (const plantel in programasEducativos[nivel][area][programa]) {
          planteles.add(plantel);
        }
      }
    }
  }
  
  return {
    totalProgramas,
    totalPlanteles: planteles.size,
    totalNiveles: Object.keys(programasEducativos).length,
    totalAreas: Object.keys(estadisticasPorArea).length,
    estadisticasPorNivel,
    estadisticasPorArea,
    planteles: Array.from(planteles).sort()
  };
}
"""
    
    return js_content

# EJECUCIÓN PRINCIPAL
print("🚀 Iniciando procesamiento del archivo Excel...")

# Cargar el archivo Excel
try:
    wb = openpyxl.load_workbook("oferta2.xlsx")
    print("✅ Archivo Excel cargado correctamente")
except FileNotFoundError:
    print("❌ Error: No se encontró el archivo 'oferta.xlsx'")
    exit(1)

# Nombres de las hojas a procesar
hojas = ["GDL"]

# Diccionario para almacenar todos los resultados
todos_resultados = defaultdict(list)

# Procesar cada hoja
for nombre_hoja in hojas:
    if nombre_hoja in wb.sheetnames:
        print(f"\n📄 Procesando hoja: {nombre_hoja}")
        ws = wb[nombre_hoja]
        resultados_hoja = procesar_hoja_mejorada(ws, nombre_hoja)
        
        # Agregar los resultados de esta hoja al diccionario general
        for plantel, programas in resultados_hoja.items():
            todos_resultados[plantel].extend(programas)
            
        print(f"✅ Hoja {nombre_hoja} procesada: {len(resultados_hoja)} planteles encontrados")
    else:
        print(f"⚠️  Hoja '{nombre_hoja}' no encontrada en el archivo")

# Generar estructura educativa
print("\n🏗️  Generando estructura educativa...")
estructura_educativa = generar_estructura_educativa(todos_resultados)

# Convertir a diccionario regular
estructura_final = {}
for nivel in sorted(estructura_educativa.keys()):
    estructura_final[nivel] = {}
    for area in sorted(estructura_educativa[nivel].keys()):
        estructura_final[nivel][area] = {}
        for programa in sorted(estructura_educativa[nivel][area].keys()):
            estructura_final[nivel][area][programa] = {}
            for plantel in sorted(estructura_educativa[nivel][area][programa].keys()):
                estructura_final[nivel][area][programa][plantel] = estructura_educativa[nivel][area][programa][plantel]

# Generar archivo JavaScript
print("\n📱 Generando archivo JavaScript...")
js_content = generar_javascript(estructura_final)

# Guardar archivo JavaScript
with open("programasEducativos.js", "w", encoding="utf-8") as archivo_js:
    archivo_js.write(js_content)

# Mostrar resumen final
print("\n📊 RESUMEN FINAL:")
print("=" * 50)

total_programas = 0
total_planteles = set()
estadisticas_por_area = defaultdict(int)

for nivel in sorted(estructura_final.keys()):
    programas_nivel = 0
    planteles_nivel = set()
    
    for area in estructura_final[nivel]:
        programas_area = len(estructura_final[nivel][area])
        programas_nivel += programas_area
        estadisticas_por_area[area] += programas_area
        
        for programa in estructura_final[nivel][area]:
            for plantel in estructura_final[nivel][area][programa]:
                planteles_nivel.add(plantel)
                total_planteles.add(plantel)
    
    total_programas += programas_nivel
    print(f"\n🎓 {nivel}: {programas_nivel} programa{'s' if programas_nivel > 1 else ''}")
    print(f"   📍 Disponible en {len(planteles_nivel)} plantel{'es' if len(planteles_nivel) > 1 else ''}")

print(f"\n📈 ESTADÍSTICAS POR ÁREA:")
for area in sorted(estadisticas_por_area.keys()):
    count = estadisticas_por_area[area]
    print(f"   🎯 {area}: {count} programa{'s' if count > 1 else ''}")

print(f"\n📈 TOTALES:")
print(f"   🎓 Total de programas únicos: {total_programas}")
print(f"   🏫 Total de planteles únicos: {len(total_planteles)}")
print(f"   📚 Total de niveles educativos: {len(estructura_final)}")
print(f"   🎯 Total de áreas: {len(estadisticas_por_area)}")

print(f"\n✅ Procesamiento completado!")
print(f"📄 Archivo generado:")
print(f"   • programasEducativos.js - Datos exportables para JavaScript")

print(f"\n🔍 EJEMPLO DE USO EN JAVASCRIPT:")
print(f"import {{ programasEducativos, obtenerProgramasPorArea, buscarProgramas }} from './programasEducativos.js';")
print(f"")
print(f"// Obtener todos los programas de Licenciaturas en Ciencias de la salud")
print(f"const programasSalud = obtenerProgramasPorArea('Ciencias de la salud', 'Licenciaturas');")
print(f"")
print(f"// Buscar programas que contengan 'administración'")
print(f"const resultados = buscarProgramas('administración');")
print(f"")
print(f"// Acceso directo a los datos")
print(f"const licenciaturas = programasEducativos.Licenciaturas;")

# Mostrar algunos ejemplos de programas encontrados
if "Licenciaturas" in estructura_final:
    print(f"\n📋 ALGUNOS PROGRAMAS DE LICENCIATURA ENCONTRADOS:")
    contador_ejemplos = 0
    for area in estructura_final["Licenciaturas"]:
        if contador_ejemplos >= 15:  # Mostrar máximo 15 ejemplos
            break
        for programa in list(estructura_final["Licenciaturas"][area].keys())[:3]:  # 3 por área
            if contador_ejemplos >= 15:
                break
            planteles_count = len(estructura_final["Licenciaturas"][area][programa])
            print(f"   • {programa} ({area}) - {planteles_count} plantel{'es' if planteles_count > 1 else ''}")
            contador_ejemplos += 1
        if contador_ejemplos < 15 and len(estructura_final["Licenciaturas"][area]) > 3:
            restantes = len(estructura_final["Licenciaturas"][area]) - 3
            print(f"   ... y {restantes} programa{'s' if restantes > 1 else ''} más en {area}")
    
    print(f"\n📊 PROGRAMAS INCLUIDOS:")
    total_licenciaturas = sum(len(estructura_final["Licenciaturas"][area]) 
                             for area in estructura_final["Licenciaturas"])
    print(f"   🎓 Total de Licenciaturas: {total_licenciaturas}")
    
    for area in sorted(estructura_final["Licenciaturas"].keys()):
        count = len(estructura_final["Licenciaturas"][area])
        print(f"   • {area}: {count} programa{'s' if count > 1 else ''}")

print(f"\n📋 NIVELES EXCLUIDOS:")
print(f"   ❌ Maestrías")
print(f"   ❌ BIS (Bachillerato Internacional)")
print(f"   ❌ Bachillerato/Preparatoria")  
print(f"   ❌ Secundaria")
print(f"   ❌ Primaria")